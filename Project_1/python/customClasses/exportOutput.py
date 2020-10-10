#!/usr/bin/env python
from openpyxl import Workbook


class TXTFiles:
    def __init__(self, ccList, incList, expDict, dataList):
        self.__ccList = ccList
        self.__incList = incList
        self.__expDict = expDict
        self.__userData = dataList
        self.__path = "./Project_1/python/output/txt_files/"

    # creating the txt file and producing the desired output
    def newFile(self, title, fileName):

        with open(self.__path + fileName, "tw") as outputFile:
            outputFile.write(title + "\n")
            outputFile.write(str((len(title) + 10) * "-") + "\n")
        return None

    # appending the data to the new files
    def appendSettings(self, fileName, list):
        with open(self.__path + fileName, "ta") as outputFile:
            if fileName == "costCentres.txt" or fileName == "income.txt":
                for item in list:
                    outputFile.write(item + "\n")

            elif fileName == "expenditure.txt":
                outputFile.write(
                    "Expenditure\tClassification\t%_Sales\t%_Distribution\t%_Production\t%Admin\tMax_Cost\n"
                )

                for item in list.items():
                    line = "{0}\t{1}\t{2}\t{3}\t{4}\t{5}\t{6}".format(
                        item[0],  # type
                        item[1][0],  # classification
                        item[1][1],  #%_Sales
                        item[1][2],  #%_Distribution
                        item[1][3],  #%_Production
                        item[1][4],  #%_Admin
                        item[1][5],  # Max_Cost
                    )
                    outputFile.write(line + "\n")

            elif fileName == "dataset.txt":
                outputFile.write(
                    "Period\tType\tCategory\tAccount\tCost Centre\tDescription\tAmount\n"
                )
                # print(list)
                for item in list:
                    outputFile.write(item + "\n")

        return None

    def createTXTfiles(self):
        self.incTypes()
        self.costCentre()
        self.expTypes()
        self.genDataSet()

        return None

    # exporting the cost centres as a .txt file
    def costCentre(self):
        self.newFile(title="Cost Centres", fileName="costCentres.txt")
        self.appendSettings(fileName="costCentres.txt", list=self.__ccList)

    # exporting the income types as a .txt file
    def incTypes(self):
        self.newFile(title="Income type", fileName="income.txt")
        self.appendSettings(fileName="income.txt", list=self.__incList)

    def expTypes(self):
        self.newFile(title="Expenditure type", fileName="expenditure.txt")
        self.appendSettings(fileName="expenditure.txt", list=self.__expDict)

    def genDataSet(self):
        self.newFile(title="Generated data", fileName="dataset.txt")
        self.appendSettings(fileName="dataset.txt", list=self.__userData)


class ExcelExport:
    def __init__(self):
        self.__path = "./Project_1/python/output/xls_files/"

    def createXLSX(self, file):
        self.xlsStructure(file)

    def xlsStructure(self, file):
        # creating the file worksheets
        wb = Workbook()
        ws = wb.active
        ws.title = "Dataset"

        ws1 = wb.create_sheet("Cost Centres")
        ws1.title = "Settings"

        # inserting headings

        ws["A1"] = "Hello World"

        # saving the file
        wb.save(self.__path + "dataset.xlsx")

